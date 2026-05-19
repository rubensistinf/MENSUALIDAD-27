from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File, Form, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from typing import List, Optional
import uvicorn
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
import uuid
import datetime
from passlib.context import CryptContext
import bcrypt as _bcrypt
from jose import JWTError, jwt
from io import BytesIO

import models
import schemas
from database import engine, get_db

models.Base.metadata.create_all(bind=engine)

# --- MIGRACION AUTOMATICA: agrega columna email si no existe ---
try:
    from sqlalchemy import text, inspect as sa_inspect
    inspector = sa_inspect(engine)
    existing_cols = [col['name'] for col in inspector.get_columns('usuarios')]
    if 'email' not in existing_cols:
        with engine.connect() as conn:
            conn.execute(text("ALTER TABLE usuarios ADD COLUMN email VARCHAR"))
            conn.commit()
        print("[Migration] Columna 'email' agregada exitosamente.")
except Exception as migration_err:
    print(f"[Migration] Nota: {migration_err}")

app = FastAPI(title="App Mensualidad Pro - UE 27 de Mayo")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- AUTH & SECURITY ---
SECRET_KEY = "super_secret_key_pro"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7 # 1 semana

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/login")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    try:
        return _bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))
    except Exception:
        return False

def get_password_hash(password: str) -> str:
    salt = _bcrypt.gensalt()
    return _bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')

def create_access_token(data: dict, expires_delta: Optional[datetime.timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.datetime.utcnow() + expires_delta
    else:
        expire = datetime.datetime.utcnow() + datetime.timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="No se pudo validar las credenciales",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
        token_data = schemas.TokenData(username=username, rol=payload.get("rol"))
    except JWTError:
        raise credentials_exception
    user = db.query(models.Usuario).filter(models.Usuario.username == token_data.username).first()
    if user is None:
        raise credentials_exception
    return user

async def get_current_active_secretaria(current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol != "secretaria" and current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No tiene permisos de secretaria")
    return current_user

async def get_current_active_admin(current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No tiene permisos de administrador")
    return current_user

@app.on_event("startup")
def create_initial_users():
    db = next(get_db())
    # --- Usuario Admin/Director ---
    admin = db.query(models.Usuario).filter(models.Usuario.username == "admin").first()
    if not admin:
        admin = models.Usuario(username="admin", rol="admin")
        db.add(admin)
    admin.email = "director@27demayo.com"
    admin.password_hash = get_password_hash("74420830")
    admin.rol = "admin"

    # --- Usuario Secretaria ---
    sec = db.query(models.Usuario).filter(models.Usuario.username == "secretaria").first()
    if not sec:
        sec = models.Usuario(username="secretaria", rol="secretaria")
        db.add(sec)
    sec.email = "secretaria@27demayo.com"
    sec.password_hash = get_password_hash("74420831")
    sec.rol = "secretaria"

    db.commit()

@app.post("/api/login", response_model=schemas.Token)
def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.Usuario).filter(models.Usuario.username == form_data.username).first()
    if not user or not verify_password(form_data.password, user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Usuario o contraseña incorrectos",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = datetime.timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username, "rol": user.rol}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer", "rol": user.rol, "username": user.username}

# --- CAMBIO DE CONTRASEÑA ---

@app.put("/api/usuarios/cambiar-password")
def cambiar_password(
    data: schemas.CambiarPassword,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user)
):
    if not verify_password(data.password_actual, current_user.password_hash):
        raise HTTPException(status_code=400, detail="Contraseña actual incorrecta")
    if len(data.password_nuevo) < 6:
        raise HTTPException(status_code=400, detail="La nueva contraseña debe tener al menos 6 caracteres")
    current_user.password_hash = get_password_hash(data.password_nuevo)
    db.commit()
    return {"message": "Contraseña cambiada exitosamente"}

# --- ESTUDIANTES ---

@app.get("/api/estudiantes", response_model=List[schemas.Estudiante])
def get_estudiantes(
    skip: int = 0, 
    limit: int = 100, 
    curso: str = None, 
    paralelo: str = None, 
    search: str = None,
    estado_pago: str = None,
    db: Session = Depends(get_db)
):
    query = db.query(models.Estudiante)
    
    if curso:
        query = query.filter(models.Estudiante.curso == curso)
    if paralelo:
        query = query.filter(models.Estudiante.paralelo == paralelo)
    if estado_pago:
        query = query.filter(models.Estudiante.estado_pago == estado_pago)
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            or_(
                models.Estudiante.nombres.ilike(search_filter),
                models.Estudiante.apellidos.ilike(search_filter),
                models.Estudiante.ci.ilike(search_filter)
            )
        )
        
    return query.offset(skip).limit(limit).all()

@app.delete("/api/estudiantes/{estudiante_id}")
def delete_estudiante(
    estudiante_id: int, 
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user)
):
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="Solo el director puede eliminar estudiantes")
    
    estudiante = db.query(models.Estudiante).filter(models.Estudiante.id == estudiante_id).first()
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
        
    db.delete(estudiante)
    db.commit()
    return {"message": "Estudiante eliminado correctamente"}

@app.delete("/api/estudiantes")
def delete_all_estudiantes(
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user)
):
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="Solo el director puede vaciar la lista de estudiantes")
        
    db.query(models.Estudiante).delete()
    db.commit()
    return {"message": "Todos los estudiantes han sido eliminados"}

@app.post("/api/estudiantes", response_model=schemas.Estudiante)
def create_estudiante(estudiante: schemas.EstudianteCreate, db: Session = Depends(get_db)):
    db_estudiante = models.Estudiante(**estudiante.model_dump())
    db.add(db_estudiante)
    db.commit()
    db.refresh(db_estudiante)
    return db_estudiante

@app.post("/api/estudiantes/upload")
async def upload_estudiantes_excel(
    file: UploadFile = File(...), 
    curso: str = Form(...), 
    paralelo: str = Form(...), 
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_active_secretaria)
):
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(contents))
        sheet = wb.active
        
        # Asumiendo que las columnas son: Nombres | Apellidos | CI (opcional)
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]:
                continue
            
            nombres = str(row[0])
            apellidos = str(row[1])
            ci = str(row[2]) if len(row) > 2 and row[2] else None
            
            db_est = models.Estudiante(nombres=nombres, apellidos=apellidos, ci=ci, curso=curso, paralelo=paralelo)
            db.add(db_est)
            count += 1
            
        db.commit()
        return {"message": f"{count} estudiantes importados exitosamente."}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error procesando archivo: {str(e)}")

@app.get("/api/estudiantes/exportar")
def exportar_estudiantes(curso: str = None, paralelo: str = None, db: Session = Depends(get_db)):
    query = db.query(models.Estudiante)
    if curso:
        query = query.filter(models.Estudiante.curso == curso)
    if paralelo:
        query = query.filter(models.Estudiante.paralelo == paralelo)
    
    estudiantes = query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inscritos"
    
    headers = ["ID", "Nombres", "Apellidos", "CI", "Curso", "Paralelo", "Estado de Pago"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill
        
    for est in estudiantes:
        ws.append([est.id, est.nombres, est.apellidos, est.ci or "-", est.curso, est.paralelo, est.estado_pago])
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 2)

    file_path = "estudiantes_export.xlsx"
    wb.save(file_path)
    
    return FileResponse(path=file_path, filename="Lista_Inscritos_27_de_Mayo.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --- PADRES Y RECIBOS ---

@app.get("/api/padres/{carnet}", response_model=schemas.Padre)
def get_padre_by_carnet(carnet: str, db: Session = Depends(get_db)):
    padre = db.query(models.Padre).filter(models.Padre.carnet == carnet).first()
    if not padre:
        raise HTTPException(status_code=404, detail="Padre no encontrado")
    return padre

@app.post("/api/padres", response_model=schemas.Padre)
def create_padre(padre: schemas.PadreCreate, db: Session = Depends(get_db)):
    db_padre = models.Padre(**padre.model_dump())
    db.add(db_padre)
    db.commit()
    db.refresh(db_padre)
    return db_padre

@app.post("/api/recibos", response_model=schemas.Recibo)
def create_recibo(
    recibo_data: schemas.ReciboCreate, 
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_active_secretaria)
):
    import random
    nro_recibo = f"REC-{random.randint(10000, 99999)}"
    
    db_recibo = models.Recibo(
        nro_recibo=nro_recibo,
        monto=recibo_data.monto,
        padre_id=recibo_data.padre_id
    )
    db.add(db_recibo)
    db.commit()
    db.refresh(db_recibo)
    
    for est_id in recibo_data.estudiantes_ids:
        estudiante = db.query(models.Estudiante).filter(models.Estudiante.id == est_id).first()
        if estudiante:
            estudiante.padre_id = recibo_data.padre_id
            estudiante.estado_pago = "Cancelado"
            
        detalle = models.DetalleRecibo(recibo_id=db_recibo.id, estudiante_id=est_id)
        db.add(detalle)
        
    # Registrar Ingreso en Caja
    ingreso = models.CajaTransaccion(
        tipo="ingreso",
        monto=recibo_data.monto,
        descripcion=f"Cobro mensualidad {nro_recibo} - Padre ID {recibo_data.padre_id}",
        usuario_id=current_user.id
    )
    db.add(ingreso)
        
    db.commit()
    db.refresh(db_recibo)
    return db_recibo

@app.post("/api/sync/recibos")
def sync_recibos(
    recibos_offline: List[schemas.ReciboCreate], 
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_active_secretaria)
):
    # Procesa múltiples recibos creados offline
    import random
    procesados = []
    for rd in recibos_offline:
        nro_recibo = f"REC-{random.randint(10000, 99999)}"
        db_recibo = models.Recibo(
            nro_recibo=nro_recibo, monto=rd.monto, padre_id=rd.padre_id
        )
        db.add(db_recibo)
        db.commit()
        db.refresh(db_recibo)
        
        for est_id in rd.estudiantes_ids:
            estudiante = db.query(models.Estudiante).filter(models.Estudiante.id == est_id).first()
            if estudiante:
                estudiante.padre_id = rd.padre_id
                estudiante.estado_pago = "Cancelado"
            detalle = models.DetalleRecibo(recibo_id=db_recibo.id, estudiante_id=est_id)
            db.add(detalle)
            
        ingreso = models.CajaTransaccion(
            tipo="ingreso",
            monto=rd.monto,
            descripcion=f"Sync Offline: Cobro mensualidad {nro_recibo} - Padre ID {rd.padre_id}",
            usuario_id=current_user.id
        )
        db.add(ingreso)
        db.commit()
        procesados.append(db_recibo.id)
    
    return {"message": f"Sincronizados {len(procesados)} recibos", "ids": procesados}

# --- CAJA ---

@app.get("/api/caja/informe", response_model=schemas.CajaInforme)
def get_caja_informe(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_active_secretaria)):
    transacciones = db.query(models.CajaTransaccion).order_by(models.CajaTransaccion.fecha.desc()).all()
    
    total_ingresos = sum(t.monto for t in transacciones if t.tipo == 'ingreso')
    total_egresos = sum(t.monto for t in transacciones if t.tipo == 'egreso')
    saldo_actual = total_ingresos - total_egresos
    
    return {
        "total_ingresos": total_ingresos,
        "total_egresos": total_egresos,
        "saldo_actual": saldo_actual,
        "transacciones": transacciones
    }

@app.post("/api/caja/egresos", response_model=schemas.CajaTransaccion)
def create_egreso(
    egreso_data: schemas.CajaTransaccionCreate, 
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_active_secretaria)
):
    if egreso_data.tipo != 'egreso':
        raise HTTPException(status_code=400, detail="El tipo debe ser 'egreso'")
        
    # Verificar si hay saldo suficiente (opcional, pero buena práctica)
    informe = get_caja_informe(db, current_user)
    if informe["saldo_actual"] < egreso_data.monto:
        raise HTTPException(status_code=400, detail="Saldo insuficiente en caja")

    db_egreso = models.CajaTransaccion(
        tipo="egreso",
        monto=egreso_data.monto,
        descripcion=egreso_data.descripcion,
        usuario_id=current_user.id
    )
    db.add(db_egreso)
    db.commit()
    db.refresh(db_egreso)
    return db_egreso

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)

# Mount the frontend directory. This MUST be at the end after all /api routes
frontend_path = os.path.join(os.path.dirname(__file__), "..", "frontend")
app.mount("/", StaticFiles(directory=frontend_path, html=True), name="frontend")

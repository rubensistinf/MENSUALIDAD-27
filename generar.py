import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import random

nombres = [
    'Juan Carlos','Maria Elena','Pedro Luis','Ana Sofia','Carlos Alberto',
    'Rosa Maria','Miguel Angel','Carmen Rosa','Luis Fernando','Gloria Patricia',
    'Jorge Enrique','Martha Cecilia','Ricardo Andres','Patricia Isabel','Fernando Jose',
    'Sandra Milena','Diego Alejandro','Claudia Marcela','Andres Felipe','Gabriela Andrea',
    'Oscar Mauricio','Monica Liliana','Hector Manuel','Adriana Paola','Roberto Carlos',
    'Liliana Marcela','Cesar Augusto','Viviana Lorena','Alejandro Jose','Diana Carolina',
    'Pablo Emilio','Laura Cristina','Rafael Antonio','Paola Andrea','Eduardo Alberto',
    'Natalia Alejandra','Rodrigo Ivan','Valentina Sofia','Sergio Andres','Daniela Fernanda',
    'Fabian Leonardo','Juliana Maria','Cristian David','Camila Alejandra','Jonathan Steven',
    'Xiomara Paola','Sebastian Felipe','Isabella Sofia','Mateo Alejandro','Valeria Daniela'
]

apellidos = [
    'Gutierrez','Mamani','Quispe','Garcia','Rodriguez',
    'Lopez','Martinez','Gonzalez','Perez','Sanchez',
    'Ramirez','Torres','Flores','Rivera','Morales',
    'Jimenez','Herrera','Medina','Vargas','Castillo',
    'Romero','Alvarez','Mendez','Ramos','Cruz',
    'Ortega','Rojas','Vega','Reyes','Diaz',
    'Soto','Aguilar','Miranda','Delgado','Fuentes',
    'Contreras','Blanco','Pacheco','Molina','Paredes',
    'Espinoza','Guerrero','Salazar','Navarro','Serrano',
    'Campos','Escobar','Ibarra','Cabrera','Ponce'
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Estudiantes'

headers = ['Nombres', 'Apellidos', 'CI']
ws.append(headers)

header_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF', size=12)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = header_fill

random.seed(42)
for i in range(50):
    n = nombres[i]
    a1 = apellidos[i % len(apellidos)]
    a2 = apellidos[(i + 7) % len(apellidos)]
    ci = str(random.randint(1000000, 9999999))
    ws.append([n, a1 + ' ' + a2, ci])

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 15

output = r'F:\ANTIGRAVITY\MENSUALIDAD-27\frontend\datos_prueba_50.xlsx'
wb.save(output)
print('OK: 50 estudiantes generados en', output)
